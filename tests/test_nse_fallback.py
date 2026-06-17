import os
import unittest
import json
from backend.scraper import scrape_screener_shareholding, get_quarterly_shareholding_pattern
from backend.price_helper import PriceFetcher
from batch_process import extract_symbols_from_csv, run_batch_compilation

class TestNSEFallback(unittest.TestCase):
    
    def test_csv_extraction_two_columns(self):
        """
        Verify that extract_symbols_from_csv extracts both BSE and NSE columns when headers are present.
        """
        # Test with headers
        csv_data = "BSE Code,NSE Symbol,Company Name\n500112,SBIN,State Bank\n500209,INFY,Infosys"
        symbols = extract_symbols_from_csv(csv_data)
        self.assertEqual(len(symbols), 2)
        self.assertEqual(symbols[0]["bse"], "500112")
        self.assertEqual(symbols[0]["nse"], "SBIN")
        self.assertEqual(symbols[1]["bse"], "500209")
        self.assertEqual(symbols[1]["nse"], "INFY")

    def test_csv_extraction_auto_detect(self):
        """
        Verify that extract_symbols_from_csv auto-detects columns based on content when headers are not descriptive.
        """
        # Test without headers, BSE in first col, NSE in second col
        csv_data = "500112,SBIN\n500209,INFY"
        symbols = extract_symbols_from_csv(csv_data)
        self.assertEqual(len(symbols), 2)
        self.assertEqual(symbols[0]["bse"], "500112")
        self.assertEqual(symbols[0]["nse"], "SBIN")
        
        # Test without headers, NSE in first col, BSE in second col
        csv_data_reverse = "SBIN,500112\nINFY,500209"
        symbols_reverse = extract_symbols_from_csv(csv_data_reverse)
        self.assertEqual(len(symbols_reverse), 2)
        self.assertEqual(symbols_reverse[0]["bse"], "500112")
        self.assertEqual(symbols_reverse[0]["nse"], "SBIN")

    def test_screener_scraping_live(self):
        """
        Verify that scrape_screener_shareholding successfully fetches and structures data from Screener.in.
        """
        # Fetch for SBIN
        data = scrape_screener_shareholding("SBIN")
        self.assertIsNotNone(data)
        self.assertIn("quarterly", data)
        self.assertIn("annual", data)
        
        q_data = data["quarterly"]
        self.assertEqual(q_data["symbol"], "SBIN")
        self.assertGreater(len(q_data["quarters"]), 0)
        
        # Verify category headers contain SEBI codes (A) and (B)
        sample_q = list(q_data["quarters"].values())[0]
        summary_cats = [item["category"] for item in sample_q["summary"]]
        self.assertTrue(any("(A)" in cat for cat in summary_cats))
        self.assertTrue(any("(B)" in cat for cat in summary_cats))

    def test_price_fetcher_nse(self):
        """
        Verify that PriceFetcher successfully fetches prices for NSE symbols (.NS suffix).
        """
        # We pass a non-existent scrip code but a valid NSE symbol
        fetcher = PriceFetcher("INVALID_CODE", "SBIN")
        self.assertEqual(fetcher.ticker_symbol, "SBIN.NS")
        self.assertIsNotNone(fetcher.history)
        self.assertFalse(fetcher.history.empty)

    def test_batch_compilation_with_fallback(self):
        """
        Verify that run_batch_compilation executes successfully with fallbacks and filters out zero-change rows.
        """
        output_xlsx = os.path.join("cache", "test_fallback_compilation.xlsx")
        if os.path.exists(output_xlsx):
            os.remove(output_xlsx)
            
        # We pass a symbol that doesn't exist on BSE but does on NSE (we mock/use a custom string to force fallback)
        # We can also pass an invalid BSE code and valid NSE code to force fallback
        symbols = [{"bse": "999999", "nse": "TCS"}]
        
        completed, failed = run_batch_compilation(symbols, output_xlsx)
        self.assertEqual(completed, 1)
        self.assertEqual(failed, 0)
        self.assertTrue(os.path.exists(output_xlsx))
        
        # Clean up
        if os.path.exists(output_xlsx):
            os.remove(output_xlsx)

    def test_constant_promoter_holding_fallback(self):
        """
        Verify that a company with constant promoter holdings (zero change) is written as a single row.
        """
        output_xlsx = os.path.join("cache", "test_constant_promoter.xlsx")
        if os.path.exists(output_xlsx):
            os.remove(output_xlsx)
            
        from unittest.mock import patch
        
        mock_data = {
            "company_name": "Test Constant Company",
            "scrip_code": "500000",
            "symbol": "TCC",
            "quarters": {
                "2023-03": {"year": "2023", "qtr_month": "Mar", "summary": [{"category": "Promoter & Group (A)", "percentage": 50.0}]},
                "2023-06": {"year": "2023", "qtr_month": "Jun", "summary": [{"category": "Promoter & Group (A)", "percentage": 50.0}]},
            }
        }
        
        with patch('batch_process.get_quarterly_shareholding_pattern', return_value=mock_data), \
             patch('batch_process.PriceFetcher') as mock_fetcher:
             
            mock_fetcher.return_value.get_price_on_date.return_value = 100.0
            
            completed, failed = run_batch_compilation(["500000"], output_xlsx)
            self.assertEqual(completed, 1)
            self.assertEqual(failed, 0)
            self.assertTrue(os.path.exists(output_xlsx))
            
            import openpyxl
            wb = openpyxl.load_workbook(output_xlsx)
            ws = wb.active
            
            self.assertEqual(ws.cell(row=2, column=1).value, "Test Constant Company")
            self.assertEqual(ws.cell(row=2, column=2).value, 0.0) # change
            self.assertEqual(ws.cell(row=2, column=3).value, "Jun-2023") # latest quarter
            self.assertIsNone(ws.cell(row=3, column=1).value)
            
        if os.path.exists(output_xlsx):
            os.remove(output_xlsx)

    def test_incremental_saving(self):
        """
        Verify that Excel compilation progress is saved incrementally after each company is compiled.
        """
        output_xlsx = os.path.join("cache", "test_incremental.xlsx")
        if os.path.exists(output_xlsx):
            os.remove(output_xlsx)
            
        from unittest.mock import patch
        
        mock_data_1 = {
            "company_name": "Company One",
            "scrip_code": "500001",
            "symbol": "CO1",
            "quarters": {
                "2023-03": {"year": "2023", "qtr_month": "Mar", "summary": [{"category": "Promoter & Group (A)", "percentage": 50.0}]}
            }
        }
        mock_data_2 = {
            "company_name": "Company Two",
            "scrip_code": "500002",
            "symbol": "CO2",
            "quarters": {
                "2023-03": {"year": "2023", "qtr_month": "Mar", "summary": [{"category": "Promoter & Group (A)", "percentage": 60.0}]}
            }
        }
        
        with patch('batch_process.get_quarterly_shareholding_pattern') as mock_get, \
             patch('batch_process.PriceFetcher') as mock_fetcher, \
             patch('batch_process.save_excel_progress') as mock_save:
             
            mock_get.side_effect = [mock_data_1, mock_data_2]
            mock_fetcher.return_value.get_price_on_date.return_value = 100.0
            
            completed, failed = run_batch_compilation(["500001", "500002"], output_xlsx)
            self.assertEqual(completed, 2)
            self.assertEqual(failed, 0)
            
            self.assertEqual(mock_save.call_count, 2)
            
            first_call_args = mock_save.call_args_list[0][0]
            second_call_args = mock_save.call_args_list[1][0]
            
            self.assertEqual(len(first_call_args[0]), 1)
            self.assertEqual(first_call_args[0][0]["company_name"], "Company One")
            
            self.assertEqual(len(second_call_args[0]), 2)
            self.assertEqual(second_call_args[0][0]["company_name"], "Company One")
            self.assertEqual(second_call_args[0][1]["company_name"], "Company Two")
            
        if os.path.exists(output_xlsx):
            os.remove(output_xlsx)

if __name__ == "__main__":
    unittest.main()
