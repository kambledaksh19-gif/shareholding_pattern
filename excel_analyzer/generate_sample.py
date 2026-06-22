import openpyxl
import os

def create_sample_excel():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Holding Changes"
    
    # Define headers
    headers = [
        "Company Name", 
        "BSE Scrip Code", 
        "Quarter", 
        "Promoter Holding %", 
        "Change in Promoter Holding",
        "CMP after 1 month",
        "CMP after 18 months"
    ]
    
    for col_idx, header in enumerate(headers, 1):
        ws.cell(row=1, column=col_idx, value=header)
        
    # Mock data rows
    # Some companies appear less than 4 times, some appear 4+ times.
    data = [
        # TATA MOTORS: 5 occurrences (should match)
        ["TATA MOTORS LTD.", "500570", "Jun-2023", "46.39%", "+0.50%", "560.20", "980.10"],
        ["TATA MOTORS LTD.", "500570", "Sep-2023", "46.89%", "+0.50%", "620.40", "1010.50"],
        ["TATA MOTORS LTD.", "500570", "Dec-2023", "46.86%", "-0.03%", "710.10", "990.20"],
        ["TATA MOTORS LTD.", "500570", "Mar-2024", "47.36%", "+0.50%", "980.50", "950.40"],
        ["TATA MOTORS LTD.", "500570", "Jun-2024", "46.86%", "-0.50%", "960.00", "940.10"],
        
        # RELIANCE: 4 occurrences (should match)
        ["RELIANCE INDUSTRIES LTD.", "500325", "Sep-2023", "50.29%", "+0.12%", "2450.00", "2850.00"],
        ["RELIANCE INDUSTRIES LTD.", "500325", "Dec-2023", "50.39%", "+0.10%", "2580.40", "2920.10"],
        ["RELIANCE INDUSTRIES LTD.", "500325", "Mar-2024", "50.35%", "-0.04%", "2900.50", "2810.00"],
        ["RELIANCE INDUSTRIES LTD.", "500325", "Jun-2024", "50.50%", "+0.15%", "2950.00", "2890.00"],
        
        # STATE BANK: 3 occurrences (should not match, count < 4)
        ["STATE BANK OF INDIA", "500112", "Jun-2023", "57.49%", "-0.10%", "580.10", "820.40"],
        ["STATE BANK OF INDIA", "500112", "Sep-2023", "57.54%", "+0.05%", "590.20", "840.10"],
        ["STATE BANK OF INDIA", "500112", "Dec-2023", "57.54%", "0.00%", "640.00", "810.50"],
        
        # INFOSYS: 4 occurrences (should match)
        ["INFOSYS LTD.", "500209", "Jun-2023", "14.94%", "-0.06%", "1310.20", "1540.20"],
        ["INFOSYS LTD.", "500209", "Sep-2023", "14.89%", "-0.05%", "1430.50", "1620.10"],
        ["INFOSYS LTD.", "500209", "Dec-2023", "14.99%", "+0.10%", "1520.10", "1590.40"],
        ["INFOSYS LTD.", "500209", "Mar-2024", "14.99%", "0.00%", "1610.40", "1550.00"],
        
        # TCS: 2 occurrences (should not match)
        ["TATA CONSULTANCY SERVICES LTD.", "500547", "Mar-2024", "72.41%", "+0.05%", "4100.00", "3900.00"],
        ["TATA CONSULTANCY SERVICES LTD.", "500547", "Jun-2024", "71.90%", "-0.51%", "3850.50", "3920.00"],
        
        # HDFC BANK: 1 occurrence (should not match)
        ["HDFC BANK LTD.", "500180", "Jun-2024", "25.60%", "+0.20%", "1620.00", "1720.00"]
    ]
    
    for r_idx, row_data in enumerate(data, 2):
        for c_idx, val in enumerate(row_data, 1):
            ws.cell(row=r_idx, column=c_idx, value=val)
            
    # Auto-adjust columns widths
    for col in ws.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        col_letter = openpyxl.utils.get_column_letter(col[0].column)
        ws.column_dimensions[col_letter].width = max(max_len + 3, 10)
        
    out_path = os.path.join(os.path.dirname(__file__), "sample_data.xlsx")
    wb.save(out_path)
    print(f"Sample Excel sheet generated successfully at: {out_path}")

if __name__ == "__main__":
    create_sample_excel()
