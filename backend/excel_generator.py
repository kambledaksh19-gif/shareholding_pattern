import os
import json
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

# Styles definitions
FONT_FAMILY = "Segoe UI"
HEADER_FILL = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid") # Dark steel blue
SECTION_FILL = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid") # Soft blue accent
TOTAL_FILL = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid") # Very light gray

HEADER_FONT = Font(name=FONT_FAMILY, size=11, bold=True, color="FFFFFF")
TITLE_FONT = Font(name=FONT_FAMILY, size=14, bold=True, color="1F4E78")
SECTION_FONT = Font(name=FONT_FAMILY, size=12, bold=True, color="000000")
BODY_FONT = Font(name=FONT_FAMILY, size=10, bold=False, color="000000")
TOTAL_FONT = Font(name=FONT_FAMILY, size=10, bold=True, color="000000")

ALIGN_LEFT = Alignment(horizontal="left", vertical="center")
ALIGN_RIGHT = Alignment(horizontal="right", vertical="center")
ALIGN_CENTER = Alignment(horizontal="center", vertical="center")

THIN_BORDER = Border(
    left=Side(style='thin', color='D9D9D9'),
    right=Side(style='thin', color='D9D9D9'),
    top=Side(style='thin', color='D9D9D9'),
    bottom=Side(style='thin', color='D9D9D9')
)

DOUBLE_BOTTOM_BORDER = Border(
    left=Side(style='thin', color='D9D9D9'),
    right=Side(style='thin', color='D9D9D9'),
    top=Side(style='thin', color='D9D9D9'),
    bottom=Side(style='double', color='000000')
)

def create_excel_workbook(data, output_path):
    """
    Generates a beautifully formatted multi-sheet Excel file.
    data: dict returned by scraper.get_shareholding_pattern
    output_path: path to save the .xlsx file
    """
    wb = openpyxl.Workbook()
    # Remove default sheet
    default_sheet = wb.active
    wb.remove(default_sheet)
    
    company_name = data.get("company_name", "Company")
    scrip_code = data.get("scrip_code", "BSE")
    years = data.get("years", {})
    
    # Process years in descending order
    for year_str in sorted(years.keys(), reverse=True):
        year_data = years[year_str]
        qtr_label = year_data.get("qtr_label", f"March {year_str}")
        
        # Sheet name limit is 31 characters
        sheet_name = f"Mar {year_str}"
        ws = wb.create_sheet(title=sheet_name)
        ws.views.sheetView[0].showGridLines = True
        
        # 1. Write Title Block
        ws.merge_cells("A1:H1")
        ws["A1"] = f"{company_name.upper()} ({scrip_code})"
        ws["A1"].font = TITLE_FONT
        ws["A1"].alignment = Alignment(horizontal="left", vertical="center")
        
        ws.merge_cells("A2:H2")
        ws["A2"] = f"Shareholding Pattern for the Annual Period ending {qtr_label}"
        ws["A2"].font = Font(name=FONT_FAMILY, size=11, italic=True, color="595959")
        ws["A2"].alignment = Alignment(horizontal="left", vertical="center")
        
        ws.row_dimensions[1].height = 28
        ws.row_dimensions[2].height = 18
        
        current_row = 4
        
        # Helper to write section headers
        def write_section_header(title):
            nonlocal current_row
            ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=8)
            cell = ws.cell(row=current_row, column=1)
            cell.value = title
            cell.font = SECTION_FONT
            cell.fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid") # Soft green fill
            cell.alignment = ALIGN_LEFT
            ws.row_dimensions[current_row].height = 24
            current_row += 2
            
        # Helper to write tables
        def write_table(headers, rows, is_promoter=False):
            nonlocal current_row
            
            # Header Row
            ws.row_dimensions[current_row].height = 24
            for col_idx, h in enumerate(headers, 1):
                cell = ws.cell(row=current_row, column=col_idx)
                cell.value = h
                cell.font = HEADER_FONT
                cell.fill = HEADER_FILL
                cell.alignment = ALIGN_CENTER
                cell.border = THIN_BORDER
                
            current_row += 1
            
            if not rows:
                ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=len(headers))
                cell = ws.cell(row=current_row, column=1)
                cell.value = "No records filed under this category."
                cell.font = Font(name=FONT_FAMILY, size=10, italic=True)
                cell.alignment = ALIGN_CENTER
                cell.border = THIN_BORDER
                current_row += 2
                return
                
            # Data Rows
            for r in rows:
                ws.row_dimensions[current_row].height = 18
                is_subtotal = r.get("shareholder_name") is None and not is_promoter
                is_total_row = "total" in str(r.get("shareholder_name") or r.get("category") or "").lower() or r.get("category") == "Grand Total"
                
                # Determine font and background
                row_font = TOTAL_FONT if (is_subtotal or is_total_row) else BODY_FONT
                row_fill = PatternFill(fill_type=None)
                if is_total_row:
                    row_fill = TOTAL_FILL
                elif is_subtotal:
                    row_fill = SECTION_FILL
                    
                row_border = DOUBLE_BOTTOM_BORDER if is_total_row else THIN_BORDER
                
                # Write cells
                for col_idx, h in enumerate(headers, 1):
                    cell = ws.cell(row=current_row, column=col_idx)
                    cell.font = row_font
                    cell.border = row_border
                    if row_fill.fill_type:
                        cell.fill = row_fill
                        
                    # Map header values
                    val = None
                    align = ALIGN_LEFT
                    num_format = None
                    
                    h_lower = h.lower()
                    if "category" in h_lower:
                        val = r.get("category") or r.get("sub_category")
                        align = ALIGN_LEFT
                    elif "name" in h_lower:
                        val = r.get("shareholder_name") or r.get("category") or ""
                        align = ALIGN_LEFT
                    elif "sub-category" in h_lower:
                        val = r.get("sub_category") or ""
                        align = ALIGN_LEFT
                    elif "shareholders" in h_lower or "no. of shareholders" in h_lower:
                        val = r.get("no_shareholders") or r.get("no_shareholder") or 0
                        align = ALIGN_RIGHT
                        num_format = "#,##0"
                    elif "pledged %" in h_lower or "pledged percentage" in h_lower or "encumbered percentage" in h_lower:
                        val = r.get("pledged_percentage") or r.get("pledge_percent") or 0.0
                        align = ALIGN_RIGHT
                        num_format = "0.00"
                    elif "pledged shares" in h_lower or "encumbered no" in h_lower:
                        val = r.get("pledged_shares") or r.get("pledge_shares") or 0
                        align = ALIGN_RIGHT
                        num_format = "#,##0"
                    elif "percentage" in h_lower or "% of total" in h_lower or "shareholding as a %" in h_lower or "percentageof_a_b_c2" in h_lower or "as a %" in h_lower:
                        val = r.get("percentage") or r.get("percent") or 0.0
                        align = ALIGN_RIGHT
                        num_format = "0.00"
                    elif "demat" in h_lower:
                        val = r.get("demat_shares") or r.get("dematerialized") or 0
                        align = ALIGN_RIGHT
                        num_format = "#,##0"
                    elif "shares held" in h_lower or "total shares" in h_lower or "fully paid" in h_lower or "shares" in h_lower:
                        val = r.get("shares") or r.get("total_shares") or 0
                        align = ALIGN_RIGHT
                        num_format = "#,##0"
                        
                    cell.value = val
                    cell.alignment = align
                    if num_format:
                        cell.number_format = num_format
                        
                current_row += 1
            current_row += 2 # spacing between tables

        # Write Table 1: Summary Statement
        write_section_header("I. SUMMARY STATEMENT OF SPECIFIED SECURITIES")
        summary_headers = ['Category of Shareholder', 'No. of Shareholders', 'Total Shares Held', 'As a % of Total Shares', 'Total Demat Shares', 'Pledged Shares', 'Pledged %']
        write_table(summary_headers, year_data.get("summary", []))
        
        # Write Table 2: Promoter Details
        write_section_header("II. STATEMENT SHOWING SHAREHOLDING PATTERN OF PROMOTER & PROMOTER GROUP")
        promoter_headers = ['Shareholder Name', 'Sub-category', 'No. of Shareholders', 'Total Shares Held', 'As a % of Total Shares', 'Total Demat Shares', 'Pledged Shares', 'Pledged %']
        write_table(promoter_headers, year_data.get("promoter", []), is_promoter=True)
        
        # Write Table 3: Public Details
        write_section_header("III. STATEMENT SHOWING SHAREHOLDING PATTERN OF PUBLIC SHAREHOLDERS")
        public_headers = ['Shareholder Name', 'Sub-category', 'No. of Shareholders', 'Total Shares Held', 'As a % of Total Shares', 'Total Demat Shares']
        write_table(public_headers, year_data.get("public", []))
        
        # Write Table 4: Non-Promoter Non-Public Details
        write_section_header("IV. STATEMENT SHOWING SHAREHOLDING PATTERN OF NON-PROMOTER NON-PUBLIC SHAREHOLDERS")
        np_headers = ['Shareholder Name', 'Sub-category', 'No. of Shareholders', 'Total Shares Held', 'As a % of Total Shares', 'Total Demat Shares']
        write_table(np_headers, year_data.get("non_promoter", []))
        
        # Auto-adjust column dimensions to avoid truncation ###
        for col in ws.columns:
            max_len = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                # Do not count Title and section rows for column widths
                if cell.row in [1, 2] or (cell.value and "statement showing" in str(cell.value).lower()):
                    continue
                if cell.value:
                    # format clean representation length
                    val_str = str(cell.value)
                    if isinstance(cell.value, (int, float)) and cell.number_format and "#,##0" in cell.number_format:
                        val_str = f"{cell.value:,.0f}"
                    max_len = max(max_len, len(val_str))
            # Set minimum width to 12 and maximum to 45
            ws.column_dimensions[col_letter].width = max(min(max_len + 3, 45), 12)
            
    # Save the file
    os.makedirs(os.path.dirname(output_path), exist_ok=True)
    wb.save(output_path)
    print(f"Excel workbook successfully created at: {output_path}")

if __name__ == "__main__":
    # Test generator using cached data
    import glob
    cache_files = glob.glob(os.path.join("cache", "*_data.json"))
    if cache_files:
        with open(cache_files[0], "r", encoding="utf-8") as f:
            test_data = json.load(f)
        create_excel_workbook(test_data, "cache/test_output.xlsx")
