import os
import csv
import io
import sys
import time
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime

# Add current directory to path to allow importing backend modules
sys.path.append(os.path.abspath(os.path.dirname(__file__)))

from backend.scraper import get_quarterly_shareholding_pattern
from backend.price_helper import PriceFetcher, get_quarter_end_date, add_months

def extract_symbols_from_csv(csv_content_or_path):
    """
    Parses CSV and returns a list of dicts: [{"bse": "...", "nse": "..."}, ...]
    Auto-detects BSE scrip code column and NSE symbol column.
    """
    rows = []
    
    try:
        if os.path.exists(csv_content_or_path):
            with open(csv_content_or_path, "r", encoding="utf-8-sig") as f:
                reader = csv.reader(f)
                rows = list(reader)
        else:
            # Treat as raw CSV string content
            f = io.StringIO(csv_content_or_path)
            reader = csv.reader(f)
            rows = list(reader)
    except Exception as e:
        print(f"Error reading CSV content: {e}")
        return []
        
    if not rows:
        return []
        
    # Clean rows of empty rows
    rows = [r for r in rows if r and any(cell.strip() for cell in r)]
    if not rows:
        return []
        
    # Simple check for header presence
    first_row = [str(x).strip().lower() for x in rows[0]]
    target_names = ["symbol", "ticker", "scripcode", "scrip code", "code", "scrip_code", "company symbol", "company code", "bse", "nse"]
    has_header = any(any(name in col for name in target_names) for col in first_row) or not any(col.isdigit() for col in first_row if col)
    
    start_row = 1 if has_header else 0
    
    bse_col_idx = None
    nse_col_idx = None
    
    # 1. Header mapping
    if has_header:
        for idx, col in enumerate(first_row):
            if "bse" in col or "scrip" in col or "code" in col:
                if "nse" not in col:
                    bse_col_idx = idx
            elif "nse" in col or "symbol" in col or "ticker" in col:
                if "bse" not in col:
                    nse_col_idx = idx
                    
    # 2. Value-based auto-detection fallback
    num_cols = len(rows[start_row]) if len(rows) > start_row else 0
    if (bse_col_idx is None or nse_col_idx is None) and num_cols > 0:
        col_scores = []
        for col in range(num_cols):
            bse_score = 0
            nse_score = 0
            for r in rows[start_row:start_row+10]:
                if col < len(r):
                    val = str(r[col]).strip()
                    if not val:
                        continue
                    if val.isdigit() and len(val) == 6:
                        bse_score += 2
                    elif val.isalpha() and 2 <= len(val) <= 10:
                        nse_score += 2
                        bse_score -= 1
                    elif val.isalnum() and 2 <= len(val) <= 10:
                        nse_score += 1
            col_scores.append({"bse": bse_score, "nse": nse_score})
            
        if bse_col_idx is None:
            best_bse = -999
            for idx, sc in enumerate(col_scores):
                if sc["bse"] > best_bse:
                    best_bse = sc["bse"]
                    bse_col_idx = idx
                    
        if nse_col_idx is None:
            best_nse = -999
            for idx, sc in enumerate(col_scores):
                if idx == bse_col_idx and num_cols > 1:
                    continue
                if sc["nse"] > best_nse:
                    best_nse = sc["nse"]
                    nse_col_idx = idx
                    
    # Fallback to defaults
    if bse_col_idx is None:
        bse_col_idx = 0
    if nse_col_idx is None:
        nse_col_idx = 1 if num_cols > 1 else 0
        
    symbols = []
    seen = set()
    for r in rows[start_row:]:
        bse_val = ""
        nse_val = ""
        if len(r) > bse_col_idx:
            bse_val = str(r[bse_col_idx]).strip()
        if len(r) > nse_col_idx:
            nse_val = str(r[nse_col_idx]).strip()
            
        if bse_val or nse_val:
            key = (bse_val, nse_val)
            if key not in seen:
                seen.add(key)
                symbols.append({
                    "bse": bse_val,
                    "nse": nse_val
                })
                
    return symbols

def save_excel_progress(all_data_rows, output_xlsx_path):
    """
    Writes the consolidated Excel sheet with the given data rows.
    Overwrites the file if it already exists.
    """
    if not all_data_rows:
        return
        
    dir_name = os.path.dirname(output_xlsx_path)
    if dir_name:
        os.makedirs(dir_name, exist_ok=True)
        
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Shareholding Performance"
    
    # Gridlines visible
    ws.views.sheetView[0].showGridLines = True
    
    # Define styles
    font_header = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    font_body = Font(name="Calibri", size=11, bold=False)
    fill_header = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    
    align_left = Alignment(horizontal="left", vertical="center")
    align_right = Alignment(horizontal="right", vertical="center")
    align_center = Alignment(horizontal="center", vertical="center")
    
    border_side = Side(style='thin', color='D9D9D9')
    border_cell = Border(left=border_side, right=border_side, top=border_side, bottom=border_side)
    
    # Write headers
    headers = [
        "company name",
        "change in promoter holding",
        "change in which quarter",
        "CMP after one month",
        "CMP after 18 months",
        "Returns(Between CMP after one month and CMP after 18 months)",
        "CMP after 24 months",
        "Returns(Between CMP after one month and CMP after 24 months)"
    ]
    
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = font_header
        cell.fill = fill_header
        cell.alignment = align_center if col_idx == 3 else (align_left if col_idx == 1 else align_right)
        cell.border = border_cell
        
    # Write data rows
    # Sort rows: Company Name alphabetical, then quarter chronological
    all_data_rows_sorted = sorted(
        all_data_rows,
        key=lambda x: (x["company_name"].lower(), x["sort_key"][0], x["sort_key"][1])
    )
    
    for row_idx, data_row in enumerate(all_data_rows_sorted, 2):
        # Company name
        c1 = ws.cell(row=row_idx, column=1, value=data_row["company_name"])
        c1.font = font_body
        c1.alignment = align_left
        c1.border = border_cell
        
        # Change in promoter holding
        change = data_row["change"]
        c2 = ws.cell(row=row_idx, column=2, value=change / 100.0)
        c2.font = font_body
        c2.alignment = align_right
        c2.border = border_cell
        c2.number_format = '+0.00%;-0.00%;0.00%'
        # Color code change text color (green for increase, red for decrease)
        if change > 0.001:
            c2.font = Font(name="Calibri", size=11, bold=False, color="385723") # Dark Green
        elif change < -0.001:
            c2.font = Font(name="Calibri", size=11, bold=False, color="C00000") # Dark Red
            
        # Change Quarter
        c3 = ws.cell(row=row_idx, column=3, value=data_row["quarter"])
        c3.font = font_body
        c3.alignment = align_center
        c3.border = border_cell
        
        # Column 4: CMP after one month
        val_cmp = data_row["cmp"]
        c4 = ws.cell(row=row_idx, column=4)
        if val_cmp is not None:
            c4.value = float(val_cmp)
            c4.number_format = '0.00'
            c4.font = font_body
        else:
            c4.value = "N/A"
            c4.font = Font(name="Calibri", size=11, italic=True, color="7F7F7F")
        c4.alignment = align_right
        c4.border = border_cell
        
        # Column 5: CMP after 18 months
        val_18 = data_row["cmp_18"]
        c5 = ws.cell(row=row_idx, column=5)
        if val_18 is not None:
            c5.value = float(val_18)
            c5.number_format = '0.00'
            c5.font = font_body
        else:
            c5.value = "N/A"
            c5.font = Font(name="Calibri", size=11, italic=True, color="7F7F7F")
        c5.alignment = align_right
        c5.border = border_cell
        
        # Column 6: Returns (Between CMP after one month and CMP after 18 months)
        c6 = ws.cell(row=row_idx, column=6)
        c6.value = f'=IF(AND(ISNUMBER(D{row_idx}), ISNUMBER(E{row_idx})), (E{row_idx}-D{row_idx})/D{row_idx}, "N/A")'
        c6.number_format = '+0.00%;-0.00%;0.00%'
        c6.font = font_body
        c6.alignment = align_right
        c6.border = border_cell
        
        # Column 7: CMP after 24 months
        val_24 = data_row["cmp_24"]
        c7 = ws.cell(row=row_idx, column=7)
        if val_24 is not None:
            c7.value = float(val_24)
            c7.number_format = '0.00'
            c7.font = font_body
        else:
            c7.value = "N/A"
            c7.font = Font(name="Calibri", size=11, italic=True, color="7F7F7F")
        c7.alignment = align_right
        c7.border = border_cell
        
        # Column 8: Returns (Between CMP after one month and CMP after 24 months)
        c8 = ws.cell(row=row_idx, column=8)
        c8.value = f'=IF(AND(ISNUMBER(D{row_idx}), ISNUMBER(G{row_idx})), (G{row_idx}-D{row_idx})/D{row_idx}, "N/A")'
        c8.number_format = '+0.00%;-0.00%;0.00%'
        c8.font = font_body
        c8.alignment = align_right
        c8.border = border_cell
        
    # Auto-size columns
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            val = str(cell.value or "")
            # Ignore formula strings for column sizing
            if val.startswith("="):
                continue
            if len(val) > max_len:
                max_len = len(val)
        ws.column_dimensions[col_letter].width = max(max_len + 4, 12)
        
    wb.save(output_xlsx_path)
    print(f"Incremental Excel progress saved at: {output_xlsx_path}")

def run_batch_compilation(symbols, output_xlsx_path, progress_callback=None):
    """
    Scrapes quarterly holdings, fetches stock prices, calculates changes, 
    and writes/saves a consolidated Excel file incrementally.
    """
    total = len(symbols)
    completed = 0
    failed = 0
    
    # Store all compiled rows across all companies
    all_data_rows = []
    
    for idx, sym in enumerate(symbols, 1):
        bse_code = ""
        nse_code = ""
        if isinstance(sym, dict):
            bse_code = sym.get("bse", "")
            nse_code = sym.get("nse", "")
        else:
            bse_code = str(sym)
            nse_code = str(sym)
            
        sym_name = bse_code if bse_code else nse_code
        print(f"[{idx}/{total}] Processing {sym_name}...")
        if progress_callback:
            progress_callback(sym, idx, total, completed, failed, "processing", None)
            
        try:
            # 1. Fetch quarterly data
            data = None
            if bse_code:
                data = get_quarterly_shareholding_pattern(bse_code)
                if not data or "error" in data:
                    print(f"BSE search failed/no-data for {bse_code}.")
                    data = None
            
            if not data and nse_code:
                print(f"Attempting to fetch via NSE code: {nse_code}...")
                data = get_quarterly_shareholding_pattern(nse_code)
                
            if not data or "error" in data:
                error_msg = data.get("error", "No data returned") if data else "No data returned"
                print(f"Error scraping {sym_name}: {error_msg}")
                failed += 1
                if progress_callback:
                    progress_callback(sym, idx, total, completed, failed, "failed", error_msg)
                continue
                
            company_name = data.get("company_name", sym_name)
            scrip_code = data.get("scrip_code", sym_name)
            symbol_ticker = data.get("symbol", "")
            
            # Override symbol_ticker if it is missing or numeric/generic and we have an NSE code from the CSV
            if (not symbol_ticker or symbol_ticker.isdigit() or "BSE_" in symbol_ticker) and nse_code:
                symbol_ticker = nse_code
            
            # 2. Sort quarters chronologically
            quarters_dict = data.get("quarters", {})
            if not quarters_dict:
                error_msg = "No quarterly data found"
                print(f"Error for {sym_name}: {error_msg}")
                failed += 1
                if progress_callback:
                    progress_callback(sym, idx, total, completed, failed, "failed", error_msg)
                continue
                
            # Helper to parse year/month for sorting
            months_order = ['Mar', 'Jun', 'Sep', 'Dec']
            quarters_sorted = sorted(
                quarters_dict.items(),
                key=lambda x: (int(x[1]["year"]), months_order.index(x[1]["qtr_month"]))
            )
            
            # 3. Instantiate PriceFetcher
            print(f"Fetching historical prices for {company_name} ({scrip_code})...")
            price_fetcher = PriceFetcher(scrip_code, symbol_ticker, company_name)
            
            # Helper to get promoter holding percentage
            def get_promoter_pct(qtr_data):
                for item in qtr_data.get("summary", []):
                    if "(A)" in item.get("category", ""):
                        return float(item.get("percentage", 0.0))
                return 0.0
            
            # 4. Calculate quarterly change and match prices
            company_rows = []
            for t, (qtr_key, qtr_data) in enumerate(quarters_sorted):
                year = qtr_data["year"]
                qtr_month = qtr_data["qtr_month"]
                promoter_pct = get_promoter_pct(qtr_data)
                
                # Calculate change from previous quarter
                if t == 0:
                    change = 0.0
                else:
                    prev_qtr_data = quarters_sorted[t-1][1]
                    prev_promoter_pct = get_promoter_pct(prev_qtr_data)
                    change = promoter_pct - prev_promoter_pct
                    
                # Format change percentage string
                change_str = f"+{change:.2f}%" if change > 0 else (f"{change:.2f}%" if change < 0 else "0.00%")
                
                # Date calculation
                quarter_end_date = get_quarter_end_date(year, qtr_month)
                cmp_price = None
                cmp_18 = None
                cmp_24 = None
                
                if quarter_end_date:
                    date_obj = datetime.strptime(quarter_end_date, "%Y-%m-%d")
                    # base date is quarter end date + 1 month (CMP after one month)
                    base_date_obj = add_months(date_obj, 1)
                    base_date_str = base_date_obj.strftime("%Y-%m-%d")
                    cmp_price = price_fetcher.get_price_on_date(base_date_str)
                    
                    # 18 months date starting from CMP after one month
                    date_18_str = add_months(base_date_obj, 18).strftime("%Y-%m-%d")
                    cmp_18 = price_fetcher.get_price_on_date(date_18_str)
                    
                    # 24 months date starting from CMP after one month
                    date_24_str = add_months(base_date_obj, 24).strftime("%Y-%m-%d")
                    cmp_24 = price_fetcher.get_price_on_date(date_24_str)
                
                # Only include the line if there was a change in promoter holdings
                if abs(change) > 0.0001:
                    company_rows.append({
                        "company_name": company_name,
                        "change": change,
                        "change_str": change_str,
                        "quarter": f"{qtr_month}-{year}",
                        "cmp": cmp_price,
                        "cmp_18": cmp_18,
                        "cmp_24": cmp_24,
                        # For sorting all rows at the end if desired (oldest first)
                        "sort_key": (int(year), months_order.index(qtr_month))
                    })
            
            # Zero-Change Fallback: if there was no change across all quarters, show the latest quarter at 0.00%
            if not company_rows and quarters_sorted:
                latest_qtr_key, latest_qtr_data = quarters_sorted[-1]
                year = latest_qtr_data["year"]
                qtr_month = latest_qtr_data["qtr_month"]
                
                quarter_end_date = get_quarter_end_date(year, qtr_month)
                cmp_price = None
                cmp_18 = None
                cmp_24 = None
                
                if quarter_end_date:
                    date_obj = datetime.strptime(quarter_end_date, "%Y-%m-%d")
                    base_date_obj = add_months(date_obj, 1)
                    base_date_str = base_date_obj.strftime("%Y-%m-%d")
                    cmp_price = price_fetcher.get_price_on_date(base_date_str)
                    
                    date_18_str = add_months(base_date_obj, 18).strftime("%Y-%m-%d")
                    cmp_18 = price_fetcher.get_price_on_date(date_18_str)
                    
                    date_24_str = add_months(base_date_obj, 24).strftime("%Y-%m-%d")
                    cmp_24 = price_fetcher.get_price_on_date(date_24_str)
                
                company_rows.append({
                    "company_name": company_name,
                    "change": 0.0,
                    "change_str": "0.00%",
                    "quarter": f"{qtr_month}-{year}",
                    "cmp": cmp_price,
                    "cmp_18": cmp_18,
                    "cmp_24": cmp_24,
                    "sort_key": (int(year), months_order.index(qtr_month))
                })
                
            all_data_rows.extend(company_rows)
            completed += 1
            
            # Save immediately (incremental progress)
            save_excel_progress(list(all_data_rows), output_xlsx_path)
            
            if progress_callback:
                progress_callback(sym, idx, total, completed, failed, "success", None)
        except Exception as e:
            error_msg = str(e)
            print(f"Exception processing {sym_name}: {error_msg}")
            failed += 1
            if progress_callback:
                progress_callback(sym, idx, total, completed, failed, "failed", error_msg)
                
        # Polite spacing between stocks in batch
        if idx < total:
            time.sleep(1.0)
            
    return completed, failed

if __name__ == "__main__":
    if len(sys.argv) < 2:
        print("Usage: python batch_process.py <csv_file_path> [output_xlsx_path]")
        sys.exit(1)
        
    csv_path = sys.argv[1]
    xlsx_path = sys.argv[2] if len(sys.argv) > 2 else "compiled_quarterly_shareholding_performance.xlsx"
    
    if not os.path.exists(csv_path):
        print(f"Error: CSV file not found at {csv_path}")
        sys.exit(1)
        
    print(f"Parsing CSV: {csv_path}...")
    symbols = extract_symbols_from_csv(csv_path)
    print(f"Found {len(symbols)} unique symbols/codes: {symbols}")
    
    if not symbols:
        print("No symbols found in CSV.")
        sys.exit(1)
        
    completed, failed = run_batch_compilation(symbols, xlsx_path)
    print(f"\nSummary: Successfully compiled {completed}/{len(symbols)} stocks. Failed: {failed}.")
